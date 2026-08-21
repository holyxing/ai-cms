"""栏目 (Category) API (P2.7)

依据: docs/17-站点树重构.md §4.1

API (以 site 为上下文):
- GET    /api/v1/sites/{site_id}/categories              树形列表 (?flat=true 退化为扁平)
- POST   /api/v1/sites/{site_id}/categories              创建
- GET    /api/v1/categories/{id}                        详情
- PATCH  /api/v1/categories/{id}                        更新 (含移动 parent)
- DELETE /api/v1/categories/{id}                        软删除 (级联子栏目)
- POST   /api/v1/categories/{id}/move                   拖拽移动 (单独端点便于前端调用)
- POST   /api/v1/categories/{id}/copy                   复制栏目 (含子栏目结构, 不复制文章) - OQ2

权限 (沿用 P1.3):
- 读: super_admin / site owner / site member
- 写: super_admin / site owner / site editor
- 软删: super_admin / site owner

设计要点:
- 物化路径 + 邻接表双保险 (沿用 Taxonomy 的成熟模式)
- 删除时级联所有后代 (path LIKE /<old_path>%)
- 移动时同步更新所有后代的 path
- 防止循环引用 (不能将栏目移到自己的后代下)
"""
import uuid
from datetime import datetime, timezone
from io import BytesIO
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, File, Query, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

import openpyxl

from app.core.deps import CurrentUser
from app.core.exceptions import BadRequest, Conflict, Forbidden, NotFound
from app.core.responses import ok
from app.db.session import get_db
from app.models.category import Category
from app.models.membership import SiteMember
from app.models.site import Site
from app.models.user import User
from app.schemas.category import (
    CategoryCreate,
    CategoryRead,
    CategoryTreeNode,
    CategoryUpdate,
)

router = APIRouter(tags=["categories"])


# === 权限 helper (复制自 taxonomies.py, 适配 Category) ===

async def _get_site_or_404(db: AsyncSession, site_id: uuid.UUID) -> Site:
    result = await db.execute(
        select(Site).where(Site.id == site_id, Site.deleted_at.is_(None))
    )
    site = result.scalar_one_or_none()
    if not site:
        raise NotFound("站点不存在")
    return site


async def _get_user_role(db: AsyncSession, site: Site, user: User) -> str | None:
    """返回 user 在 site 的角色: 'owner' | 'editor' | 'viewer' | None (无访问)"""
    if user.is_super_admin:
        return "owner"
    if site.owner_id == user.id:
        return "owner"
    result = await db.execute(
        select(SiteMember.name).where(
            SiteMember.site_id == site.id,
            SiteMember.user_id == user.id,
            SiteMember.deleted_at.is_(None),
        )
    )
    return result.scalar_one_or_none()


def _can_read(role: str | None) -> bool:
    return role in ("owner", "editor", "viewer")


def _can_write(role: str | None) -> bool:
    return role in ("owner", "editor")


def _can_delete(role: str | None) -> bool:
    return role == "owner"


# === 物化路径维护 ===

def _make_path(parent_path: Optional[str], self_id: uuid.UUID) -> str:
    """构造物化路径: 根 /<id>/, 子 /<parent_path><id>/"""
    if parent_path:
        return f"{parent_path}{self_id}/"
    return f"/{self_id}/"


async def _update_descendant_paths(
    db: AsyncSession, old_path: str, new_path: str
) -> None:
    """节点移动后, 更新所有后代的 path (前缀替换)"""
    if old_path == new_path:
        return
    result = await db.execute(
        select(Category).where(
            Category.path.like(f"{old_path}%"),
            Category.path != old_path,
            Category.deleted_at.is_(None),
        )
    )
    for c in result.scalars().all():
        c.path = new_path + c.path[len(old_path):]
    await db.flush()


async def _is_descendant(
    db: AsyncSession, ancestor_id: uuid.UUID, candidate_id: uuid.UUID
) -> bool:
    """判断 candidate 是否 ancestor 的后代"""
    anc = await db.get(Category, ancestor_id)
    if not anc:
        return False
    cand = await db.get(Category, candidate_id)
    if not cand:
        return False
    return cand.path.startswith(anc.path) and cand.id != anc.id


# === 树构建 ===

def _build_tree(items: list[Category]) -> list[dict]:
    """从扁平列表构造树 (path 已排序)"""
    nodes = {
        c.id: {
            "id": str(c.id),
            "site_id": str(c.site_id),
            "parent_id": str(c.parent_id) if c.parent_id else None,
            "name": c.name,
            "slug": c.slug,
            "path": c.path,
            "description": c.description,
            "order_num": c.order_num,
            "seo": c.seo or {},
            "content_count": c.content_count,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
            "children": [],
        }
        for c in items
    }
    roots: list[dict] = []
    for c in items:
        node = nodes[c.id]
        if c.parent_id and c.parent_id in nodes:
            nodes[c.parent_id]["children"].append(node)
        else:
            roots.append(node)
    return roots


def _to_read_dict(c: Category) -> dict:
    return {
        "id": str(c.id),
        "site_id": str(c.site_id),
        "parent_id": str(c.parent_id) if c.parent_id else None,
        "name": c.name,
        "slug": c.slug,
        "path": c.path,
        "description": c.description,
        "order_num": c.order_num,
        "seo": c.seo or {},
        "content_count": c.content_count,
        "template": c.template or "default",
        "content_template": c.content_template or "default",
        "created_at": c.created_at.isoformat() if c.created_at else None,
        "updated_at": c.updated_at.isoformat() if c.updated_at else None,
    }


# === Excel 批量导入 (栏目) helper ===

def _slugify(name: str) -> str:
    """生成合法 slug：英文保留；中文转全拼；其余字符转连字符，限长 64。"""
    import re
    from pypinyin import Style, lazy_pinyin

    raw = str(name or "").strip().lower()
    pinyin = "".join(lazy_pinyin(raw, style=Style.NORMAL))
    s = re.sub(r'[^a-z0-9-]+', '-', pinyin)
    s = re.sub(r'-+', '-', s).strip('-')
    return s[:64] or 'cat'


def _unique_slug(base: str, existing: set[str]) -> str:
    """在 existing 集合内找唯一 slug, 冲突加 -2 / -3 后缀"""
    if base not in existing:
        return base
    i = 2
    while f"{base}-{i}" in existing:
        i += 1
    return f"{base}-{i}"


# === 端点 ===

@router.get("/sites/{site_id}/categories", response_model=None)
async def list_categories(
    site_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
    flat: bool = Query(False, description="true 返扁平, false 返树形 (默认)"),
):
    """列出站点的栏目树 (默认树形)"""
    site = await _get_site_or_404(db, site_id)
    role = await _get_user_role(db, site, current_user)
    if not _can_read(role):
        raise Forbidden("无权访问该站点")

    # 一次拉全部 (栏目数 < 200 没问题, 大站再加 limit/pagination)
    # 注意：树渲染顺序必须按同级 order_num，而不能按 path，
    # 否则拖拽排序写入后，前端看到的顺序仍会被 path 覆盖。
    result = await db.execute(
        select(Category)
        .where(Category.site_id == site_id, Category.deleted_at.is_(None))
        .order_by(Category.parent_id, Category.order_num, Category.created_at, Category.id)
    )
    items = result.scalars().all()

    if flat:
        return ok({
            "site_id": str(site_id),
            "total": len(items),
            "items": [_to_read_dict(c) for c in items],
        })

    return ok({
        "site_id": str(site_id),
        "total": len(items),
        "tree": _build_tree(items),
    })


@router.post("/sites/{site_id}/categories", response_model=None, status_code=201)
async def create_category(
    site_id: uuid.UUID,
    body: CategoryCreate,
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """创建栏目"""
    site = await _get_site_or_404(db, site_id)
    role = await _get_user_role(db, site, current_user)
    if not _can_write(role):
        raise Forbidden("无权创建栏目")

    # 校验 parent_id
    parent_path: Optional[str] = None
    if body.parent_id:
        parent = await db.get(Category, body.parent_id)
        if not parent or parent.site_id != site_id or parent.deleted_at:
            raise BadRequest("父栏目不存在或不属于该站点")
        parent_path = parent.path

    # 同级最大 order_num + 10
    max_order = (await db.execute(
        select(func.coalesce(func.max(Category.order_num), 0)).where(
            Category.site_id == site_id,
            Category.parent_id == body.parent_id,
            Category.deleted_at.is_(None),
        )
    )).scalar() or 0

    c = Category(
        site_id=site_id,
        parent_id=body.parent_id,
        name=body.name,
        slug=body.slug,
        path="/",  # flush 后填
        description=body.description,
        order_num=max_order + 10,
        template=body.template or "default",
        content_template=body.content_template or "default",
    )
    db.add(c)
    try:
        await db.flush()  # 拿 id
        c.path = _make_path(parent_path, c.id)
        await db.commit()
    except IntegrityError as e:
        await db.rollback()
        msg = str(e.orig) if e.orig else str(e)
        if "uq_categories_site_slug" in msg or "slug" in msg:
            raise Conflict(f"slug '{body.slug}' 在该站点下已存在", code=40901) from e
        raise
    await db.refresh(c)
    return ok(_to_read_dict(c), message="栏目已创建")


@router.post("/sites/{site_id}/categories/import", response_model=None, status_code=201)
async def import_categories(
    site_id: uuid.UUID,
    file: Annotated[UploadFile, File(description="xlsx 文件, 3 列: 一级栏目 / 二级栏目 / 三级栏目")],
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """批量导入栏目 (Excel .xlsx)

    Excel 格式约定:
    - 必须有 3 列, 按列名识别: 一级栏目 / 二级栏目 / 三级栏目
    - 空单元格表示"继承上一行的上一级" (例: 一级=产品中心, 二级=海石, 三级=(空) → 产品中心→海石)
    - 单次最多导入 100 个栏目
    - slug 重名自动加 -2 / -3 后缀 (不报错)
    - 默认栏目 (站点的现有栏目) 不受影响, 不重复创建

    限制: 原子事务 — 任何一行失败全部回滚
    """
    site = await _get_site_or_404(db, site_id)
    role = await _get_user_role(db, site, current_user)
    if not _can_write(role):
        raise Forbidden("无权导入栏目")

    content = await file.read()
    if not content:
        raise BadRequest("文件为空")

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise BadRequest(f"Excel 文件解析失败: {e}")

    ws = wb.active
    if ws is None:
        raise BadRequest("Excel 文件没有活动工作表")

    # 找列索引 (按列名识别: 一级栏目 / 二级栏目 / 三级栏目)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise BadRequest("Excel 文件缺少表头")

    col_idx: dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        h_str = str(h).strip()
        if "一级" in h_str:
            col_idx["level1"] = i
        elif "二级" in h_str:
            col_idx["level2"] = i
        elif "三级" in h_str:
            col_idx["level3"] = i

    if "level1" not in col_idx:
        raise BadRequest("Excel 必须有「一级栏目」列")

    # 解析行 → (level, name) 列表 (空单元格跳过, 非空才追加)
    raw_rows: list[tuple[int, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        for key in ("level1", "level2", "level3"):
            idx = col_idx.get(key)
            if idx is None or idx >= len(row):
                continue
            v = row[idx]
            if v is None or str(v).strip() == "":
                continue
            raw_rows.append((int(key[-1]), str(v).strip()))

    if not raw_rows:
        raise BadRequest("Excel 中没有有效数据")

    if len(raw_rows) > 100:
        raise BadRequest(f"单次最多导入 100 个栏目, 当前 {len(raw_rows)} 个")

    # 预拉 site 下已存在的 slugs (避免冲突)
    existing_q = await db.execute(
        select(Category.slug).where(
            Category.site_id == site_id,
            Category.deleted_at.is_(None),
        )
    )
    existing_slugs = set(existing_q.scalars().all())

    # 栈跟踪当前 parent: [{level, id, path}]
    parent_stack: list[dict] = []
    created: list[dict] = []

    try:
        for level, name in raw_rows:
            # 弹出比当前 level 深的父级 (新的一级或二级来临)
            while parent_stack and parent_stack[-1]["level"] >= level:
                parent_stack.pop()

            parent = parent_stack[-1] if parent_stack else None
            parent_id = parent["id"] if parent else None
            parent_path = parent["path"] if parent else "/"

            base_slug = _slugify(name)
            slug = _unique_slug(base_slug, existing_slugs)
            existing_slugs.add(slug)

            max_order = (await db.execute(
                select(func.coalesce(func.max(Category.order_num), 0)).where(
                    Category.site_id == site_id,
                    Category.parent_id == parent_id,
                    Category.deleted_at.is_(None),
                )
            )).scalar() or 0

            c = Category(
                site_id=site_id,
                parent_id=parent_id,
                name=name,
                slug=slug,
                path="/",  # flush 后填
                order_num=max_order + 10,
                template="default",
                content_template="default",
            )
            db.add(c)
            await db.flush()  # 拿 id
            c.path = _make_path(parent_path, c.id)

            created.append({
                "id": str(c.id),
                "name": c.name,
                "slug": slug,
                "level": level,
                "parent_id": str(c.parent_id) if c.parent_id else None,
            })
            parent_stack.append({"level": level, "id": c.id, "path": c.path})

        await db.commit()
    except IntegrityError as e:
        await db.rollback()
        msg = str(e.orig) if e.orig else str(e)
        raise Conflict(f"数据库约束冲突: {msg}", code=40901) from e
    except Exception:
        await db.rollback()
        raise

    return ok({
        "created": created,
        "total": len(created),
    }, message=f"成功导入 {len(created)} 个栏目")


@router.get("/categories/{category_id}", response_model=None)
async def get_category(
    category_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """栏目详情 (含子栏目数)"""
    c = await db.get(Category, category_id)
    if not c or c.deleted_at:
        raise NotFound("栏目不存在")

    site = await _get_site_or_404(db, c.site_id)
    role = await _get_user_role(db, site, current_user)
    if not _can_read(role):
        raise Forbidden("无权访问")

    children_count = (await db.execute(
        select(Category.id).where(
            Category.parent_id == category_id,
            Category.deleted_at.is_(None),
        )
    )).scalars().all()

    data = _to_read_dict(c)
    data["children_count"] = len(children_count)
    return ok(data)


@router.patch("/categories/{category_id}", response_model=None)
async def update_category(
    category_id: uuid.UUID,
    body: CategoryUpdate,
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """更新栏目 (含移动 parent)"""
    c = await db.get(Category, category_id)
    if not c or c.deleted_at:
        raise NotFound("栏目不存在")

    site = await _get_site_or_404(db, c.site_id)
    role = await _get_user_role(db, site, current_user)
    if not _can_write(role):
        raise Forbidden("无权修改")

    # 移动 parent? (用 model_fields_set 区分 "没传" 和 "传了 null")
    if "parent_id" in body.model_fields_set:
        new_parent_id = body.parent_id
        if new_parent_id == c.id:
            raise BadRequest("不能将自己设为父栏目")
        if new_parent_id:
            if await _is_descendant(db, c.id, new_parent_id):
                raise BadRequest("不能将栏目移动到自己的后代下")
            new_parent = await db.get(Category, new_parent_id)
            if not new_parent or new_parent.site_id != c.site_id or new_parent.deleted_at:
                raise BadRequest("父栏目不存在")

        old_path = c.path
        if new_parent_id:
            c.parent_id = new_parent_id
            c.path = _make_path((await db.get(Category, new_parent_id)).path, c.id)
        else:
            c.parent_id = None
            c.path = _make_path(None, c.id)
        await _update_descendant_paths(db, old_path, c.path)

    if body.name is not None:
        c.name = body.name
    if body.slug is not None:
        c.slug = body.slug
    if body.description is not None:
        c.description = body.description
    if body.order_num is not None:
        c.order_num = body.order_num
    if body.seo is not None:
        c.seo = body.seo
    if body.template is not None:
        c.template = body.template
    if body.content_template is not None:
        c.content_template = body.content_template

    try:
        await db.commit()
    except IntegrityError as e:
        await db.rollback()
        msg = str(e.orig) if e.orig else str(e)
        if "uq_categories_site_slug" in msg or "slug" in msg:
            raise Conflict(f"slug '{body.slug}' 已被占用", code=40901) from e
        raise
    await db.refresh(c)
    return ok(_to_read_dict(c), message="已更新")


@router.delete("/categories/{category_id}", response_model=None)
async def delete_category(
    category_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """软删除栏目 (级联所有子栏目)"""
    c = await db.get(Category, category_id)
    if not c or c.deleted_at:
        raise NotFound("栏目不存在")

    site = await _get_site_or_404(db, c.site_id)
    role = await _get_user_role(db, site, current_user)
    if not _can_delete(role):
        raise Forbidden("仅 owner 可删除栏目")

    # 级联软删所有后代
    descendants = (await db.execute(
        select(Category).where(
            Category.path.like(f"{c.path}%"),
            Category.deleted_at.is_(None),
        )
    )).scalars().all()

    now = datetime.now(timezone.utc)
    for d in descendants:
        d.deleted_at = now
    # 把属于这些栏目的 content.category_id 置 NULL (SET NULL 不会触发, 软删是 UPDATE)
    from sqlalchemy import update
    from app.models.content import Content
    await db.execute(
        update(Content)
        .where(Content.category_id.in_([d.id for d in descendants]))
        .values(category_id=None)
    )
    await db.commit()
    return ok(message=f"已删除栏目及 {len(descendants) - 1} 个子栏目")


@router.delete("/categories/{category_id}/index-page", response_model=None)
async def delete_category_index_page(
    category_id: uuid.UUID,
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """删除栏目已发布的 index.html 静态文件"""
    c = await db.get(Category, category_id)
    if not c or c.deleted_at:
        raise NotFound("栏目不存在")
    site = await _get_site_or_404(db, c.site_id)
    role = await _get_user_role(db, site, current_user)
    if not _can_delete(role):
        raise Forbidden("仅 owner 可删除")

    # 构建栏目相对路径（含父栏目）
    cat_path = c.slug
    if c.parent_id:
        parent = await db.get(Category, c.parent_id)
        if parent and parent.slug:
            cat_path = f"{parent.slug}/{c.slug}"

    from app.services.static_cleanup import delete_category_index
    result = delete_category_index(site.slug, cat_path)
    return ok(result, message="已删除栏目首页" if result["removed"] else "栏目首页不存在")


# === 移动 (拖拽) ===

class MovePayload(BaseModel):
    parent_id: Optional[uuid.UUID] = Field(None, description="目标父栏目, null = 根")
    position: int = Field(0, ge=0, description="在目标父下的位置 (0=最前)")


@router.post("/categories/{category_id}/move", response_model=None)
async def move_category(
    category_id: uuid.UUID,
    body: MovePayload,
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """拖拽移动：支持改父级，也支持同级排序。"""
    c = await db.get(Category, category_id)
    if not c or c.deleted_at:
        raise NotFound("栏目不存在")
    site = await _get_site_or_404(db, c.site_id)
    role = await _get_user_role(db, site, current_user)
    if not _can_write(role):
        raise Forbidden("无权移动")

    if body.parent_id == c.id:
        raise BadRequest("不能将自己设为父栏目")
    if body.parent_id and await _is_descendant(db, c.id, body.parent_id):
        raise BadRequest("不能将栏目移动到自己的后代下")

    # 算新 path
    old_path = c.path
    if body.parent_id:
        new_parent = await db.get(Category, body.parent_id)
        if not new_parent or new_parent.site_id != c.site_id or new_parent.deleted_at:
            raise BadRequest("父栏目不存在")
        c.parent_id = body.parent_id
        c.path = _make_path(new_parent.path, c.id)
    else:
        c.parent_id = None
        c.path = _make_path(None, c.id)
    await _update_descendant_paths(db, old_path, c.path)

    # 在目标父栏目下按 position 重排同级 order_num
    siblings = (await db.execute(
        select(Category).where(
            Category.site_id == c.site_id,
            Category.parent_id == c.parent_id,
            Category.id != c.id,
            Category.deleted_at.is_(None),
        ).order_by(Category.order_num, Category.id)
    )).scalars().all()
    pos = min(max(body.position, 0), len(siblings))
    siblings.insert(pos, c)
    for idx, item in enumerate(siblings, start=1):
        item.order_num = idx * 10

    await db.commit()
    await db.refresh(c)
    return ok(_to_read_dict(c), message="已移动")


# === 复制 (OQ2) ===

class CopyPayload(BaseModel):
    name_suffix: str = Field("-copy", description="复制后名称后缀")
    slug_suffix: str = Field("-copy", description="复制后 slug 后缀 (重复时自动加序号)")


@router.post("/categories/{category_id}/copy", response_model=None, status_code=201)
async def copy_category(
    category_id: uuid.UUID,
    body: CopyPayload,
    current_user: CurrentUser,
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """复制栏目 (含子栏目结构, 不复制文章)

    递归:
    1. 复制当前节点, name/slug 加后缀 (重复自动加序号)
    2. 复制所有子栏目到新节点下
    3. 维护 path
    """
    src = await db.get(Category, category_id)
    if not src or src.deleted_at:
        raise NotFound("栏目不存在")
    site = await _get_site_or_404(db, src.site_id)
    role = await _get_user_role(db, site, current_user)
    if not _can_write(role):
        raise Forbidden("无权复制")

    # 找不冲突的 slug
    new_slug = f"{src.slug}{body.slug_suffix}"
    counter = 1
    while True:
        existing = (await db.execute(
            select(Category).where(
                Category.site_id == src.site_id,
                Category.slug == new_slug,
                Category.deleted_at.is_(None),
            )
        )).scalar_one_or_none()
        if not existing:
            break
        counter += 1
        new_slug = f"{src.slug}{body.slug_suffix}{counter}"

    # 复制当前节点 (新 id, 同 parent, name/slug 改)
    new_id = uuid.uuid4()
    new_node = Category(
        id=new_id,
        site_id=src.site_id,
        parent_id=src.parent_id,
        name=f"{src.name}{body.name_suffix}" if counter == 1 else f"{src.name}{body.name_suffix}{counter}",
        slug=new_slug,
        path=_make_path(
            (await db.get(Category, src.parent_id)).path if src.parent_id else None,
            new_id,
        ),
        description=src.description,
        order_num=src.order_num + 1,  # 排到原栏目后面
        seo=src.seo or {},
        content_count=0,  # 复制的栏目没有文章
        template=src.template,
        content_template=src.content_template,
    )
    db.add(new_node)
    await db.flush()

    # 递归复制子栏目
    await _copy_descendants(db, src.id, new_id, src.site_id, body)

    await db.commit()
    await db.refresh(new_node)
    return ok(_to_read_dict(new_node), message="栏目已复制")


async def _copy_descendants(
    db: AsyncSession, src_parent_id: uuid.UUID, dst_parent_id: uuid.UUID,
    site_id: uuid.UUID, body: CopyPayload,
) -> None:
    """递归复制 src_parent_id 的所有子栏目到 dst_parent_id 下"""
    children = (await db.execute(
        select(Category).where(
            Category.parent_id == src_parent_id,
            Category.deleted_at.is_(None),
        ).order_by(Category.order_num)
    )).scalars().all()

    for child in children:
        # 找不冲突的 slug
        new_slug = f"{child.slug}{body.slug_suffix}"
        counter = 1
        while True:
            existing = (await db.execute(
                select(Category).where(
                    Category.site_id == site_id,
                    Category.slug == new_slug,
                    Category.deleted_at.is_(None),
                )
            )).scalar_one_or_none()
            if not existing:
                break
            counter += 1
            new_slug = f"{child.slug}{body.slug_suffix}{counter}"

        new_id = uuid.uuid4()
        new_node = Category(
            id=new_id,
            site_id=site_id,
            parent_id=dst_parent_id,
            name=f"{child.name}{body.name_suffix}" if counter == 1 else f"{child.name}{body.name_suffix}{counter}",
            slug=new_slug,
            path=_make_path(
                (await db.get(Category, dst_parent_id)).path,
                new_id,
            ),
            description=child.description,
            order_num=child.order_num,
            seo=child.seo or {},
            content_count=0,
            template=child.template,
            content_template=child.content_template,
        )
        db.add(new_node)
        await db.flush()

        # 递归
        await _copy_descendants(db, child.id, new_id, site_id, body)
